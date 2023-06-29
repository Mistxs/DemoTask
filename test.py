import sqlite3
import random
from datetime import date

import openpyxl
from faker import Faker
import logging


fake = Faker()

logging.basicConfig(filename='test.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

dbase = 'test.db'

def createTable(db):
    conn = sqlite3.connect(db)
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS Users 
                      (userId INTEGER PRIMARY KEY, 
                       age INTEGER
                       )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS Items (
                          itemId INTEGER PRIMARY KEY,
                          price INTEGER
                    )''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS Purchases 
                      (   purchaseId INTEGER PRIMARY KEY,
                          userId INTEGER,
                          itemId INTEGER,
                          date DATE,
                          FOREIGN KEY (userId) REFERENCES Users(userId),
                          FOREIGN KEY (itemId) REFERENCES Items(itemId)
                       )''')
    conn.commit()
    conn.close()
def dropTable(db):
    logging.info(f"Start dropTable to f{db}")
    try:
        conn = sqlite3.connect(db)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE Users")
        cursor.execute("DROP TABLE Items")
        cursor.execute("DROP TABLE Purchases")
        conn.commit()
        conn.close()
        print("DROP TABLES SUCCESFULLLY")
    except sqlite3.OperationalError as e:
            logging.error(f"Ошибка удаления таблицы: {e}")
def insertData(db):
    logging.info(f"Start insertData to f{db}")
    conn = sqlite3.connect(db)
    cursor = conn.cursor()
    for i in range(1, 100):
        try:
            id = random.randint(1, 10000000)
            age = random.randint(14, 90)
            queryToUsers = "INSERT INTO Users (userId, age) VALUES (?,?)"
            valuesToUsers = (id, age)
            cursor.execute(queryToUsers, valuesToUsers)

            itemId = random.randint(10000000, 99999999)
            price = random.randint(1, 200000)
            queryToItems = "INSERT INTO Items (itemId, price) VALUES (?,?)"
            valuesToItems = (itemId, price)
            cursor.execute(queryToItems, valuesToItems)

            purchaseId = random.randint(10000000, 99999999)
            start_date = date(2021, 1, 1)
            random_date = fake.date_between_dates(date_start=start_date, date_end=date.today())

            queryToPurchases = "INSERT INTO Purchases (purchaseId, userId, itemId, date) VALUES (?,?,?,?)"
            valuesToPurchases = (purchaseId,id,itemId,random_date)
            cursor.execute(queryToPurchases, valuesToPurchases)

        except sqlite3.IntegrityError as e:
            logging.error(f"Ошибка вставки данных: {e}")



    conn.commit()
    conn.close()
    print("INSERT TABLES SUCCESFULLLY")
def safeResult1():
    conn = sqlite3.connect(dbase)
    cursor = conn.cursor()

    query = '''
        SELECT month,
               strftime('%Y', p.date) AS year,
               AVG(CASE WHEN u.age BETWEEN 18 AND 25 THEN i.price END) AS 'Лица 18-25',
               AVG(CASE WHEN u.age BETWEEN 26 AND 35 THEN i.price END) AS 'Лица 26-35'
        FROM (
            SELECT strftime('%m', p.date) AS month, p.date
            FROM Purchases p
            JOIN Users u ON u.userId = p.userId
            WHERE u.age BETWEEN 18 AND 25
        
            UNION
        
            SELECT strftime('%m', p.date) AS month, p.date
            FROM Purchases p
            JOIN Users u ON u.userId = p.userId
            WHERE u.age BETWEEN 26 AND 35
        ) AS p
        JOIN Purchases pu ON pu.date = p.date
        JOIN Items i ON i.itemId = pu.itemId
        JOIN Users u ON u.userId = pu.userId
        GROUP BY 1, 2
        ORDER BY 2, 1;
    
    '''
    cursor.execute(query)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Месяц', 'Год', 'Лица 18-25', 'Лица 26-35']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(cursor, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)
    workbook.save("results1.xlsx")
def safeResult2():
    conn = sqlite3.connect(dbase)
    cursor = conn.cursor()

    query = '''
    SELECT strftime('%m', p.date) as month, strftime('%Y', p.date) as year, sum(i.price)
    FROM Purchases p
    JOIN Users u ON u.userId = p.userId
    join Items I on I.itemId = p.itemId
    WHERE u.age > 35
    group by 1, 2
    order by 3 desc

    '''
    cursor.execute(query)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Месяц', 'Год', 'Сумма']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(cursor, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)
    workbook.save("results2.xlsx")
def safeResult3():
    conn = sqlite3.connect(dbase)
    cursor = conn.cursor()

    query = '''
    SELECT i.itemId as Good, i.price as Price, count(p.itemId) as amount,
           i.price * COUNT(p.itemId) as TotalCost
    FROM Purchases p
    INNER JOIN Items i ON p.itemId = i.itemId
    WHERE strftime('%Y',p.date) = '2023'
    GROUP BY 1
    ORDER BY 4 DESC
    LIMIT 1

    '''
    cursor.execute(query)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['ID Позиции', 'Цена', 'Количество', 'Сумма']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(cursor, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)
    workbook.save("results3.xlsx")
def safeResult4():
    conn = sqlite3.connect(dbase)
    cursor = conn.cursor()

    query = '''
    SELECT
    t.itemId as Good,
    t.sum as GoodsSum,
    r.totalSum as TotalSum,
    round((t.sum * 100.0 / r.totalSum),2) as Ratio
    FROM (
    SELECT
        i.itemId,
        SUM(i.price) as sum
    FROM Purchases p
    JOIN Items i ON p.itemId = i.itemId
    WHERE strftime('%Y', p.date) = '2023'
    GROUP BY 1
    ORDER BY 2 DESC
    LIMIT 3
    ) AS t
    CROSS JOIN (
        SELECT SUM(i.price) as totalSum
        FROM Purchases p
        JOIN Items i ON p.itemId = i.itemId
        WHERE strftime('%Y', p.date) = '2023'
    ) AS r

    '''
    cursor.execute(query)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Товар', 'Выручка', 'Общая выручка', 'Соотношение']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(cursor, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)
    workbook.save("results4.xlsx")
def job():
    dropTable(dbase)
    createTable(dbase)
    insertData(dbase)
    safeResult1()
    safeResult2()
    safeResult3()
    safeResult4()

job()

